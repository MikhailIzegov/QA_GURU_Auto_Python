import os
import xlrd
from openpyxl import load_workbook

from hw_work_with_files.conftest import TEMP_PATH

# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_work_with_xlsx():
    workbook = load_workbook(os.path.join(TEMP_PATH, 'import_ou_xlsx.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'OU001', 'Значение этой ячейки неверно'

    # но если мы удалим, тест не воспроизведем снова, потому что мы не создаем в коде этот файл заново
    # os.remove(os.path.join(TEMP_PATH, 'import_ou_xlsx.xlsx'))

