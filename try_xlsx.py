# Нужно предустановить pip3 install openpyxl
from openpyxl import load_workbook

workbook = load_workbook('resource/import_ou_xlsx.xlsx')
sheet = workbook.active
print(sheet.cell(row=3, column=2).value)
